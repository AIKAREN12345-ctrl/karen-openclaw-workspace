import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart, LineChart
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# Load files
df_original = pd.read_excel('C:/Users/Karen/.openclaw/media/inbound/file_191---1d6a3fcf-e32b-413e-8746-b239ed12319a.xlsx')
df_cleaned = pd.read_excel('C:/Users/Karen/.openclaw/media/inbound/file_192---d557f616-4976-476b-ae78-fb9a252cc8fd.xlsx')

# Clean data for analysis
df_analysis = df_cleaned[df_cleaned['Sales Person'] != 'Sales Person'].copy()
df_analysis['Date'] = pd.to_datetime(df_analysis['Date'])
df_analysis['Month'] = df_analysis['Date'].dt.strftime('%Y-%m')
df_analysis['Amount'] = df_analysis['Amount'].replace('[\$,]', '', regex=True).astype(float)
df_analysis['Boxes Shipped'] = pd.to_numeric(df_analysis['Boxes Shipped'], errors='coerce')

# Create workbook
wb = Workbook()

# Styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14, color="1F4E78")
subtitle_font = Font(bold=True, size=12, color="2E75B5")
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# ==================== SHEET 1: Original Data (File 191) ====================
ws1 = wb.active
ws1.title = "Original Data (Raw)"

# Write all data as-is from original file
for r_idx, row in enumerate(dataframe_to_rows(df_original, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
        cell.border = border

# Auto-adjust columns
for col in ws1.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 30)
    ws1.column_dimensions[column].width = adjusted_width

# ==================== SHEET 2: Cleaned Data (File 192) ====================
ws2 = wb.create_sheet("Cleaned Data")

for r_idx, row in enumerate(dataframe_to_rows(df_cleaned, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
        cell.border = border

for col in ws2.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 30)
    ws2.column_dimensions[column].width = adjusted_width

# ==================== SHEET 3: Workings (Pivot Tables) ====================
ws3 = wb.create_sheet("Workings - Pivot Tables")

ws3['A1'] = "WORKINGS - Pivot Tables and Calculations"
ws3['A1'].font = title_font
ws3.merge_cells('A1:F1')

# Pivot 1: Country Summary
ws3['A3'] = "Pivot Table 1: Sales by Country"
ws3['A3'].font = subtitle_font
country_pivot = df_analysis.groupby('Country').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum',
    'Sales Person': 'nunique'
}).reset_index()
country_pivot.columns = ['Country', 'Sum of Amount', 'Sum of Boxes', 'Count of Sales People']

row = 4
for col_idx, header in enumerate(country_pivot.columns, 1):
    cell = ws3.cell(row=row, column=col_idx, value=header)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = border

row = 5
for _, data_row in country_pivot.iterrows():
    for col_idx, value in enumerate(data_row, 1):
        cell = ws3.cell(row=row, column=col_idx, value=value)
        cell.border = border
        if col_idx == 2:
            cell.number_format = '#,##0'
    row += 1

# Pivot 2: Product Summary
ws3[f'A{row+2}'] = "Pivot Table 2: Sales by Product"
ws3[f'A{row+2}'].font = subtitle_font
row += 3

product_pivot = df_analysis.groupby('Product').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum'
}).reset_index().sort_values('Amount', ascending=False)
product_pivot.columns = ['Product', 'Sum of Amount', 'Sum of Boxes']

for col_idx, header in enumerate(product_pivot.columns, 1):
    cell = ws3.cell(row=row, column=col_idx, value=header)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = border

row += 1
for _, data_row in product_pivot.head(10).iterrows():
    for col_idx, value in enumerate(data_row, 1):
        cell = ws3.cell(row=row, column=col_idx, value=value)
        cell.border = border
        if col_idx == 2:
            cell.number_format = '#,##0'
    row += 1

# Pivot 3: Monthly Summary
ws3[f'A{row+2}'] = "Pivot Table 3: Sales by Month"
ws3[f'A{row+2}'].font = subtitle_font
row += 3

monthly_pivot = df_analysis.groupby('Month').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum'
}).reset_index()
monthly_pivot.columns = ['Month', 'Sum of Amount', 'Sum of Boxes']

for col_idx, header in enumerate(monthly_pivot.columns, 1):
    cell = ws3.cell(row=row, column=col_idx, value=header)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = border

row += 1
for _, data_row in monthly_pivot.iterrows():
    for col_idx, value in enumerate(data_row, 1):
        cell = ws3.cell(row=row, column=col_idx, value=value)
        cell.border = border
        if col_idx == 2:
            cell.number_format = '#,##0'
    row += 1

for col in ['A', 'B', 'C', 'D']:
    ws3.column_dimensions[col].width = 20

# ==================== Q1: Sales by Country with Chart ====================
ws_q1 = wb.create_sheet("Q1 - Country Analysis")

ws_q1['A1'] = "Q1: Sales Performance by Country"
ws_q1['A1'].font = title_font
ws_q1.merge_cells('A1:G1')
ws_q1['A1'].alignment = Alignment(horizontal='center')

# Headers
headers = ['Country', 'Total Revenue (€)', 'Total Boxes', 'Sales People', 'Avg Revenue/Box', '% of Total']
for col_idx, header in enumerate(headers, 1):
    cell = ws_q1.cell(row=3, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# Data
total_rev = df_analysis['Amount'].sum()
country_data = df_analysis.groupby('Country').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum',
    'Sales Person': 'nunique'
}).reset_index().sort_values('Amount', ascending=False)

row = 4
chart_data_start = 4
for _, data_row in country_data.iterrows():
    ws_q1.cell(row=row, column=1, value=data_row['Country']).border = border
    ws_q1.cell(row=row, column=2, value=data_row['Amount']).border = border
    ws_q1.cell(row=row, column=2).number_format = '#,##0'
    ws_q1.cell(row=row, column=3, value=data_row['Boxes Shipped']).border = border
    ws_q1.cell(row=row, column=4, value=data_row['Sales Person']).border = border
    ws_q1.cell(row=row, column=5, value=f"=B{row}/C{row}").border = border
    ws_q1.cell(row=row, column=6, value=f"=B{row}/{total_rev}").border = border
    ws_q1.cell(row=row, column=6).number_format = '0.0%'
    for col in range(1, 7):
        ws_q1.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    row += 1

chart_data_end = row - 1

# Add Bar Chart
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Revenue by Country"
chart1.y_axis.title = 'Revenue (€)'
chart1.x_axis.title = 'Country'

data = Reference(ws_q1, min_col=2, min_row=3, max_row=chart_data_end, max_col=2)
cats = Reference(ws_q1, min_col=1, min_row=4, max_row=chart_data_end)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
chart1.width = 15
chart1.height = 10

ws_q1.add_chart(chart1, "I3")

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_q1.column_dimensions[col].width = 18

# ==================== Q2: Top Products with Chart ====================
ws_q2 = wb.create_sheet("Q2 - Product Analysis")

ws_q2['A1'] = "Q2: Top 5 Products Analysis"
ws_q2['A1'].font = title_font
ws_q2.merge_cells('A1:F1')
ws_q2['A1'].alignment = Alignment(horizontal='center')

# Top 5 by Revenue
ws_q2['A3'] = "Top 5 Products by Revenue"
ws_q2['A3'].font = subtitle_font
ws_q2.merge_cells('A3:E3')

headers = ['Rank', 'Product', 'Revenue (€)', 'Boxes', '% of Total']
for col_idx, header in enumerate(headers, 1):
    cell = ws_q2.cell(row=4, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

product_rev = df_analysis.groupby('Product').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum'
}).reset_index().sort_values('Amount', ascending=False).head(5)

row = 5
for idx, data_row in product_rev.iterrows():
    ws_q2.cell(row=row, column=1, value=row-4).border = border
    ws_q2.cell(row=row, column=2, value=data_row['Product']).border = border
    ws_q2.cell(row=row, column=3, value=data_row['Amount']).border = border
    ws_q2.cell(row=row, column=3).number_format = '#,##0'
    ws_q2.cell(row=row, column=4, value=data_row['Boxes Shipped']).border = border
    ws_q2.cell(row=row, column=5, value=f"=C{row}/{total_rev}").border = border
    ws_q2.cell(row=row, column=5).number_format = '0.0%'
    for col in range(1, 6):
        ws_q2.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    row += 1

# Add Pie Chart for Top 5
pie = PieChart()
labels = Reference(ws_q2, min_col=2, min_row=5, max_row=9)
data = Reference(ws_q2, min_col=3, min_row=4, max_row=9)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Top 5 Products by Revenue"
pie.width = 12
pie.height = 10

ws_q2.add_chart(pie, "G3")

for col in ['A', 'B', 'C', 'D', 'E']:
    ws_q2.column_dimensions[col].width = 20

# ==================== Q3: Sales Person Performance ====================
ws_q3 = wb.create_sheet("Q3 - Sales People")

ws_q3['A1'] = "Q3: Top 10 Sales People Performance"
ws_q3['A1'].font = title_font
ws_q3.merge_cells('A1:G1')
ws_q3['A1'].alignment = Alignment(horizontal='center')

headers = ['Rank', 'Sales Person', 'Revenue (€)', 'Boxes', 'Countries', 'Products']
for col_idx, header in enumerate(headers, 1):
    cell = ws_q3.cell(row=3, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

sales_perf = df_analysis.groupby('Sales Person').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum',
    'Country': 'nunique',
    'Product': 'nunique'
}).reset_index().sort_values('Amount', ascending=False).head(10)

row = 4
for idx, data_row in sales_perf.iterrows():
    ws_q3.cell(row=row, column=1, value=row-3).border = border
    ws_q3.cell(row=row, column=2, value=data_row['Sales Person']).border = border
    ws_q3.cell(row=row, column=3, value=data_row['Amount']).border = border
    ws_q3.cell(row=row, column=3).number_format = '#,##0'
    ws_q3.cell(row=row, column=4, value=data_row['Boxes Shipped']).border = border
    ws_q3.cell(row=row, column=5, value=data_row['Country']).border = border
    ws_q3.cell(row=row, column=6, value=data_row['Product']).border = border
    for col in range(1, 7):
        ws_q3.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    row += 1

# Bar chart for top 10
chart3 = BarChart()
chart3.type = "bar"
chart3.style = 11
chart3.title = "Top 10 Sales People by Revenue"
chart3.y_axis.title = 'Sales Person'
chart3.x_axis.title = 'Revenue (€)'

data = Reference(ws_q3, min_col=3, min_row=3, max_row=13, max_col=3)
cats = Reference(ws_q3, min_col=2, min_row=4, max_row=13)
chart3.add_data(data, titles_from_data=True)
chart3.set_categories(cats)
chart3.width = 15
chart3.height = 12

ws_q3.add_chart(chart3, "I3")

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_q3.column_dimensions[col].width = 18

# ==================== Q4: Monthly Trends with Line Chart ====================
ws_q4 = wb.create_sheet("Q4 - Monthly Trends")

ws_q4['A1'] = "Q4: Monthly Sales Trends (Jan-Aug 2022)"
ws_q4['A1'].font = title_font
ws_q4.merge_cells('A1:G1')
ws_q4['A1'].alignment = Alignment(horizontal='center')

headers = ['Month', 'Revenue (€)', 'Boxes', 'MoM Rev %', 'MoM Box %', '% of Total']
for col_idx, header in enumerate(headers, 1):
    cell = ws_q4.cell(row=3, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

monthly = df_analysis.groupby('Month').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum'
}).reset_index().sort_values('Month')

row = 4
prev_row = None
for idx, data_row in monthly.iterrows():
    ws_q4.cell(row=row, column=1, value=data_row['Month']).border = border
    ws_q4.cell(row=row, column=2, value=data_row['Amount']).border = border
    ws_q4.cell(row=row, column=2).number_format = '#,##0'
    ws_q4.cell(row=row, column=3, value=data_row['Boxes Shipped']).border = border
    ws_q4.cell(row=row, column=6, value=f"=B{row}/{total_rev}").border = border
    ws_q4.cell(row=row, column=6).number_format = '0.0%'
    
    if prev_row:
        ws_q4.cell(row=row, column=4, value=f"=(B{row}-B{prev_row})/B{prev_row}").border = border
        ws_q4.cell(row=row, column=4).number_format = '+0.0%;-0.0%'
        ws_q4.cell(row=row, column=5, value=f"=(C{row}-C{prev_row})/C{prev_row}").border = border
        ws_q4.cell(row=row, column=5).number_format = '+0.0%;-0.0%'
    else:
        ws_q4.cell(row=row, column=4, value="N/A").border = border
        ws_q4.cell(row=row, column=5, value="N/A").border = border
    
    for col in range(1, 7):
        ws_q4.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    prev_row = row
    row += 1

# Line chart
line_chart = LineChart()
line_chart.title = "Monthly Revenue Trend"
line_chart.style = 12
line_chart.y_axis.title = 'Revenue (€)'
line_chart.x_axis.title = 'Month'

data = Reference(ws_q4, min_col=2, min_row=3, max_row=row-1, max_col=2)
cats = Reference(ws_q4, min_col=1, min_row=4, max_row=row-1)
line_chart.add_data(data, titles_from_data=True)
line_chart.set_categories(cats)
line_chart.width = 15
line_chart.height = 10

ws_q4.add_chart(line_chart, "I3")

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_q4.column_dimensions[col].width = 15

# ==================== Q5: Product-Country Matrix ====================
ws_q5 = wb.create_sheet("Q5 - Product-Country Matrix")

ws_q5['A1'] = "Q5: Product Performance by Country (Revenue Matrix)"
ws_q5['A1'].font = title_font
ws_q5.merge_cells('A1:H1')
ws_q5['A1'].alignment = Alignment(horizontal='center')

matrix = df_analysis.pivot_table(values='Amount', index='Product', columns='Country', aggfunc='sum', fill_value=0)
matrix['Total'] = matrix.sum(axis=1)
matrix = matrix.sort_values('Total', ascending=False)

countries = list(matrix.columns)
ws_q5.cell(row=3, column=1, value="Product").fill = header_fill
ws_q5.cell(row=3, column=1).font = header_font
ws_q5.cell(row=3, column=1).border = border
ws_q5.cell(row=3, column=1).alignment = Alignment(horizontal='center')

for col_idx, country in enumerate(countries, 2):
    cell = ws_q5.cell(row=3, column=col_idx, value=country)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

row = 4
for product, data_row in matrix.iterrows():
    ws_q5.cell(row=row, column=1, value=product).border = border
    for col_idx, country in enumerate(countries, 2):
        cell = ws_q5.cell(row=row, column=col_idx, value=data_row[country])
        cell.number_format = '#,##0'
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 1

ws_q5.column_dimensions['A'].width = 25
for i in range(2, len(countries) + 2):
    ws_q5.column_dimensions[get_column_letter(i)].width = 15

# Save
output_path = 'C:/Users/Karen/.openclaw/workspace/Chocolate_Sales_Analysis_Complete.xlsx'
wb.save(output_path)
print(f"Complete Excel file created: {output_path}")
print("\nSheets:")
print("1. Original Data (Raw) - untouched from file 191")
print("2. Cleaned Data - untouched from file 192")
print("3. Workings - Pivot Tables showing calculations")
print("4. Q1 - Country Analysis with Bar Chart")
print("5. Q2 - Product Analysis with Pie Chart")
print("6. Q3 - Sales People with Bar Chart")
print("7. Q4 - Monthly Trends with Line Chart")
print("8. Q5 - Product-Country Matrix")
