import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# Load and clean data
df = pd.read_excel('C:/Users/Karen/.openclaw/media/inbound/file_191---1d6a3fcf-e32b-413e-8746-b239ed12319a.xlsx')
df = df[df['Sales Person'] != 'Sales Person'].reset_index(drop=True)
df = df[['Sales Person', 'Country', 'Product', 'Date', 'Amount', 'Boxes Shipped']]
df['Date'] = pd.to_datetime(df['Date'])
df['Month'] = df['Date'].dt.strftime('%Y-%m')
df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
df['Boxes Shipped'] = pd.to_numeric(df['Boxes Shipped'], errors='coerce')

# Create workbook
wb = Workbook()

# Keep the first sheet as Q1 and rename it
ws1 = wb.active
ws1.title = "Q1"

# Styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ==================== Q1: Sales by Country ====================
ws1['A1'] = "Q1: Sales Performance by Country"
ws1['A1'].font = title_font
ws1.merge_cells('A1:F1')
ws1['A1'].alignment = Alignment(horizontal='center')

ws1['A3'] = "Country"
ws1['B3'] = "Total Revenue (€)"
ws1['C3'] = "Total Boxes"
ws1['D3'] = "Sales People"
ws1['E3'] = "Avg Revenue per Box (€)"
ws1['F3'] = "% of Total Revenue"

# Header formatting
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    cell = ws1[f'{col}3']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# Country analysis
country_sales = df.groupby('Country').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum',
    'Sales Person': 'nunique'
}).reset_index()
country_sales.columns = ['Country', 'Total Revenue', 'Total Boxes', 'Sales People']
country_sales = country_sales.sort_values('Total Revenue', ascending=False)

total_revenue = df['Amount'].sum()

row = 4
for _, data_row in country_sales.iterrows():
    ws1[f'A{row}'] = data_row['Country']
    ws1[f'B{row}'] = data_row['Total Revenue']
    ws1[f'C{row}'] = data_row['Total Boxes']
    ws1[f'D{row}'] = data_row['Sales People']
    ws1[f'E{row}'] = f"=B{row}/C{row}"
    ws1[f'F{row}'] = f"=B{row}/{total_revenue}"
    ws1[f'F{row}'].number_format = '0.0%'
    ws1[f'B{row}'].number_format = '#,##0'
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws1[f'{col}{row}'].border = border
        ws1[f'{col}{row}'].alignment = Alignment(horizontal='center')
    row += 1

# Adjust column widths
ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 25
ws1.column_dimensions['F'].width = 20

# ==================== Q2: Top Products ====================
ws2 = wb.create_sheet("Q2")

ws2['A1'] = "Q2: Top 5 Products Analysis"
ws2['A1'].font = title_font
ws2.merge_cells('A1:E1')
ws2['A1'].alignment = Alignment(horizontal='center')

# Top 5 by Revenue
ws2['A3'] = "Top 5 Products by Revenue"
ws2['A3'].font = Font(bold=True, size=12)
ws2.merge_cells('A3:E3')

ws2['A4'] = "Rank"
ws2['B4'] = "Product"
ws2['C4'] = "Total Revenue (€)"
ws2['D4'] = "Total Boxes"
ws2['E4'] = "% of Total Revenue"

for col in ['A', 'B', 'C', 'D', 'E']:
    cell = ws2[f'{col}4']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

product_rev = df.groupby('Product').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum'
}).reset_index()
product_rev = product_rev.sort_values('Amount', ascending=False).head(5)

row = 5
rank = 1
for _, data_row in product_rev.iterrows():
    ws2[f'A{row}'] = rank
    ws2[f'B{row}'] = data_row['Product']
    ws2[f'C{row}'] = data_row['Amount']
    ws2[f'D{row}'] = data_row['Boxes Shipped']
    ws2[f'E{row}'] = f"=C{row}/{total_revenue}"
    ws2[f'E{row}'].number_format = '0.0%'
    ws2[f'C{row}'].number_format = '#,##0'
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws2[f'{col}{row}'].border = border
        ws2[f'{col}{row}'].alignment = Alignment(horizontal='center')
    row += 1
    rank += 1

# Top 5 by Volume
ws2['A11'] = "Top 5 Products by Volume (Boxes)"
ws2['A11'].font = Font(bold=True, size=12)
ws2.merge_cells('A11:E11')

ws2['A12'] = "Rank"
ws2['B12'] = "Product"
ws2['C12'] = "Total Boxes"
ws2['D12'] = "Total Revenue (€)"
ws2['E12'] = "Avg Revenue per Box"

for col in ['A', 'B', 'C', 'D', 'E']:
    cell = ws2[f'{col}12']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

product_vol = df.groupby('Product').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum'
}).reset_index()
product_vol = product_vol.sort_values('Boxes Shipped', ascending=False).head(5)

row = 13
rank = 1
for _, data_row in product_vol.iterrows():
    ws2[f'A{row}'] = rank
    ws2[f'B{row}'] = data_row['Product']
    ws2[f'C{row}'] = data_row['Boxes Shipped']
    ws2[f'D{row}'] = data_row['Amount']
    ws2[f'E{row}'] = f"=D{row}/C{row}"
    ws2[f'D{row}'].number_format = '#,##0'
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws2[f'{col}{row}'].border = border
        ws2[f'{col}{row}'].alignment = Alignment(horizontal='center')
    row += 1
    rank += 1

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 20
ws2.column_dimensions['E'].width = 22

# ==================== Q3: Sales Person Performance ====================
ws3 = wb.create_sheet("Q3")

ws3['A1'] = "Q3: Top 10 Sales People Performance"
ws3['A1'].font = title_font
ws3.merge_cells('A1:F1')
ws3['A1'].alignment = Alignment(horizontal='center')

ws3['A3'] = "Rank"
ws3['B3'] = "Sales Person"
ws3['C3'] = "Total Revenue (€)"
ws3['D3'] = "Total Boxes"
ws3['E3'] = "Countries Covered"
ws3['F3'] = "Products Sold"

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    cell = ws3[f'{col}3']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

sales_perf = df.groupby('Sales Person').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum',
    'Country': 'nunique',
    'Product': 'nunique'
}).reset_index()
sales_perf.columns = ['Sales Person', 'Total Revenue', 'Total Boxes', 'Countries', 'Products']
sales_perf = sales_perf.sort_values('Total Revenue', ascending=False).head(10)

row = 4
rank = 1
for _, data_row in sales_perf.iterrows():
    ws3[f'A{row}'] = rank
    ws3[f'B{row}'] = data_row['Sales Person']
    ws3[f'C{row}'] = data_row['Total Revenue']
    ws3[f'D{row}'] = data_row['Total Boxes']
    ws3[f'E{row}'] = data_row['Countries']
    ws3[f'F{row}'] = data_row['Products']
    ws3[f'C{row}'].number_format = '#,##0'
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws3[f'{col}{row}'].border = border
        ws3[f'{col}{row}'].alignment = Alignment(horizontal='center')
    row += 1
    rank += 1

ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 22
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 15
ws3.column_dimensions['E'].width = 18
ws3.column_dimensions['F'].width = 15

# ==================== Q4: Monthly Trends ====================
ws4 = wb.create_sheet("Q4")

ws4['A1'] = "Q4: Monthly Sales Trends (Jan-Aug 2022)"
ws4['A1'].font = title_font
ws4.merge_cells('A1:F1')
ws4['A1'].alignment = Alignment(horizontal='center')

ws4['A3'] = "Month"
ws4['B3'] = "Revenue (€)"
ws4['C3'] = "Boxes Shipped"
ws4['D3'] = "MoM Revenue Change"
ws4['E3'] = "MoM Boxes Change"
ws4['F3'] = "% of Total Revenue"

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    cell = ws4[f'{col}3']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

monthly = df.groupby('Month').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum'
}).reset_index()
monthly = monthly.sort_values('Month')

row = 4
prev_row = None
for _, data_row in monthly.iterrows():
    ws4[f'A{row}'] = data_row['Month']
    ws4[f'B{row}'] = data_row['Amount']
    ws4[f'C{row}'] = data_row['Boxes Shipped']
    ws4[f'F{row}'] = f"=B{row}/{total_revenue}"
    ws4[f'F{row}'].number_format = '0.0%'
    ws4[f'B{row}'].number_format = '#,##0'
    
    if prev_row:
        ws4[f'D{row}'] = f"=(B{row}-B{prev_row})/B{prev_row}"
        ws4[f'E{row}'] = f"=(C{row}-C{prev_row})/C{prev_row}"
        ws4[f'D{row}'].number_format = '+0.0%;-0.0%'
        ws4[f'E{row}'].number_format = '+0.0%;-0.0%'
    else:
        ws4[f'D{row}'] = "N/A"
        ws4[f'E{row}'] = "N/A"
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws4[f'{col}{row}'].border = border
        ws4[f'{col}{row}'].alignment = Alignment(horizontal='center')
    prev_row = row
    row += 1

ws4.column_dimensions['A'].width = 12
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 18
ws4.column_dimensions['D'].width = 22
ws4.column_dimensions['E'].width = 20
ws4.column_dimensions['F'].width = 20

# ==================== Q5: Product-Country Matrix ====================
ws5 = wb.create_sheet("Q5")

ws5['A1'] = "Q5: Product Performance by Country (Revenue Matrix)"
ws5['A1'].font = title_font
ws5.merge_cells('A1:H1')
ws5['A1'].alignment = Alignment(horizontal='center')

# Create pivot matrix
matrix = df.pivot_table(values='Amount', index='Product', columns='Country', aggfunc='sum', fill_value=0)
matrix['Total'] = matrix.sum(axis=1)
matrix = matrix.sort_values('Total', ascending=False)

# Header
ws5['A3'] = "Product"
ws5['A3'].fill = header_fill
ws5['A3'].font = header_font
ws5['A3'].alignment = Alignment(horizontal='center')
ws5['A3'].border = border

col_idx = 2
countries = list(matrix.columns)
for country in countries:
    col_letter = chr(64 + col_idx)
    ws5[f'{col_letter}3'] = country
    ws5[f'{col_letter}3'].fill = header_fill
    ws5[f'{col_letter}3'].font = header_font
    ws5[f'{col_letter}3'].alignment = Alignment(horizontal='center')
    ws5[f'{col_letter}3'].border = border
    col_idx += 1

# Data
row = 4
for product, data_row in matrix.iterrows():
    ws5[f'A{row}'] = product
    ws5[f'A{row}'].border = border
    ws5[f'A{row}'].alignment = Alignment(horizontal='left')
    
    col_idx = 2
    for country in countries:
        col_letter = chr(64 + col_idx)
        ws5[f'{col_letter}{row}'] = data_row[country]
        ws5[f'{col_letter}{row}'].number_format = '#,##0'
        ws5[f'{col_letter}{row}'].border = border
        ws5[f'{col_letter}{row}'].alignment = Alignment(horizontal='center')
        col_idx += 1
    row += 1

ws5.column_dimensions['A'].width = 25
for i in range(2, col_idx):
    col_letter = chr(64 + i)
    ws5.column_dimensions[col_letter].width = 15

# Save workbook
output_path = 'C:/Users/Karen/.openclaw/workspace/Chocolate_Sales_Analysis.xlsx'
wb.save(output_path)
print(f"Excel file created: {output_path}")
print("\nSheets created with formulas and formatting:")
print("- Q1: Sales by Country (with formulas for avg revenue per box and % of total)")
print("- Q2: Top Products (revenue and volume rankings)")
print("- Q3: Sales Person Performance (top 10 with metrics)")
print("- Q4: Monthly Trends (with MoM change formulas)")
print("- Q5: Product-Country Matrix (pivot table)")
