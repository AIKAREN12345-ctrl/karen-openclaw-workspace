import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# Load and clean data
df = pd.read_excel('C:/Users/Karen/.openclaw/media/inbound/file_191---1d6a3fcf-e32b-413e-8746-b239ed12319a.xlsx')
df = df[df['Sales Person'] != 'Sales Person'].reset_index(drop=True)
df = df[['Sales Person', 'Country', 'Product', 'Date', 'Amount', 'Boxes Shipped']]
df['Date'] = pd.to_datetime(df['Date'])
df['Month'] = df['Date'].dt.strftime('%Y-%m')
df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
df['Boxes Shipped'] = pd.to_numeric(df['Boxes Shipped'], errors='coerce')

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# ==================== Q1: Sales by Country ====================
ws1 = wb.create_sheet("Q1")
ws1.append(["Q1: Sales Performance by Country"])
ws1.append([])

# Country analysis
country_sales = df.groupby('Country').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum',
    'Sales Person': 'nunique'
}).reset_index()
country_sales.columns = ['Country', 'Total Revenue (€)', 'Total Boxes', 'Sales People']
country_sales = country_sales.sort_values('Total Revenue (€)', ascending=False)

ws1.append(['Country', 'Total Revenue (€)', 'Total Boxes', 'Sales People', 'Avg Revenue per Box'])
for _, row in country_sales.iterrows():
    avg_rev = row['Total Revenue (€)'] / row['Total Boxes'] if row['Total Boxes'] > 0 else 0
    ws1.append([row['Country'], row['Total Revenue (€)'], row['Total Boxes'], row['Sales People'], round(avg_rev, 2)])

# ==================== Q2: Top Products ====================
ws2 = wb.create_sheet("Q2")
ws2.append(["Q2: Top 5 Products by Revenue and Volume"])
ws2.append([])

# Top 5 by revenue
ws2.append(["Top 5 Products by Revenue"])
ws2.append(['Product', 'Total Revenue (€)', 'Total Boxes', '% of Total Revenue'])
product_rev = df.groupby('Product')['Amount'].sum().sort_values(ascending=False).head(5)
total_rev = df['Amount'].sum()
for product in product_rev.index:
    rev = product_rev[product]
    boxes = df[df['Product'] == product]['Boxes Shipped'].sum()
    pct = (rev / total_rev) * 100
    ws2.append([product, rev, boxes, f"{pct:.1f}%"])

ws2.append([])
ws2.append(["Top 5 Products by Volume (Boxes)"])
ws2.append(['Product', 'Total Boxes', 'Total Revenue (€)'])
product_vol = df.groupby('Product')['Boxes Shipped'].sum().sort_values(ascending=False).head(5)
for product in product_vol.index:
    boxes = product_vol[product]
    rev = df[df['Product'] == product]['Amount'].sum()
    ws2.append([product, boxes, rev])

# ==================== Q3: Sales Person Performance ====================
ws3 = wb.create_sheet("Q3")
ws3.append(["Q3: Top 10 Sales People by Revenue"])
ws3.append([])

sales_perf = df.groupby('Sales Person').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum',
    'Country': 'nunique',
    'Product': 'nunique'
}).reset_index()
sales_perf.columns = ['Sales Person', 'Total Revenue (€)', 'Total Boxes', 'Countries', 'Products']
sales_perf = sales_perf.sort_values('Total Revenue (€)', ascending=False).head(10)

ws3.append(['Rank', 'Sales Person', 'Total Revenue (€)', 'Total Boxes', 'Countries Covered', 'Products Sold'])
for i, (_, row) in enumerate(sales_perf.iterrows(), 1):
    ws3.append([i, row['Sales Person'], row['Total Revenue (€)'], row['Total Boxes'], row['Countries'], row['Products']])

# ==================== Q4: Monthly Trends ====================
ws4 = wb.create_sheet("Q4")
ws4.append(["Q4: Monthly Sales Trends (Jan-Aug 2022)"])
ws4.append([])

monthly = df.groupby('Month').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum'
}).reset_index()
monthly.columns = ['Month', 'Revenue (€)', 'Boxes Shipped']

ws4.append(['Month', 'Revenue (€)', 'Boxes Shipped', 'MoM Revenue Change', 'MoM Boxes Change'])
prev_rev = None
prev_boxes = None
for _, row in monthly.iterrows():
    rev_change = ""
    boxes_change = ""
    if prev_rev is not None:
        rev_pct = ((row['Revenue (€)'] - prev_rev) / prev_rev) * 100
        boxes_pct = ((row['Boxes Shipped'] - prev_boxes) / prev_boxes) * 100
        rev_change = f"{rev_pct:+.1f}%"
        boxes_change = f"{boxes_pct:+.1f}%"
    ws4.append([row['Month'], row['Revenue (€)'], row['Boxes Shipped'], rev_change, boxes_change])
    prev_rev = row['Revenue (€)']
    prev_boxes = row['Boxes Shipped']

# ==================== Q5: Product-Country Matrix ====================
ws5 = wb.create_sheet("Q5")
ws5.append(["Q5: Product Performance by Country (Revenue Matrix)"])
ws5.append([])

# Create pivot matrix
matrix = df.pivot_table(values='Amount', index='Product', columns='Country', aggfunc='sum', fill_value=0)
matrix['Total'] = matrix.sum(axis=1)
matrix = matrix.sort_values('Total', ascending=False)

# Write header
header = ['Product'] + list(matrix.columns)
ws5.append(header)

# Write data
for product, row in matrix.iterrows():
    ws5.append([product] + list(row.values))

# Save workbook
output_path = 'C:/Users/Karen/.openclaw/workspace/Chocolate_Sales_Analysis.xlsx'
wb.save(output_path)
print(f"Excel file created: {output_path}")
print("\nSheets created:")
print("- Q1: Sales by Country")
print("- Q2: Top Products")
print("- Q3: Sales Person Performance")
print("- Q4: Monthly Trends")
print("- Q5: Product-Country Matrix")
