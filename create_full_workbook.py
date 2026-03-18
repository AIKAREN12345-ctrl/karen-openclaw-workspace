import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Load and clean data
df = pd.read_excel('C:/Users/Karen/.openclaw/media/inbound/file_191---1d6a3fcf-e32b-413e-8746-b239ed12319a.xlsx')
df = df[df['Sales Person'] != 'Sales Person'].reset_index(drop=True)
df = df[['Sales Person', 'Country', 'Product', 'Date', 'Amount', 'Boxes Shipped']]
df['Date'] = pd.to_datetime(df['Date'])
df['Month'] = df['Date'].dt.strftime('%Y-%m')
df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
df['Boxes Shipped'] = pd.to_numeric(df['Boxes Shipped'], errors='coerce')

# Create workbook
wb = Workbook()

# ==================== ORIGINAL DATA SHEET ====================
ws_data = wb.active
ws_data.title = "Chocolate Sales Data"

# Styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title
ws_data['A1'] = "Chocolate Sales Data - Original Dataset"
ws_data['A1'].font = title_font
ws_data.merge_cells('A1:F1')
ws_data['A1'].alignment = Alignment(horizontal='center')

# Headers
headers = ['Sales Person', 'Country', 'Product', 'Date', 'Amount', 'Boxes Shipped']
for col_idx, header in enumerate(headers, 1):
    cell = ws_data.cell(row=3, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# Data
for row_idx, row_data in enumerate(df.values, 4):
    for col_idx, value in enumerate(row_data[:6], 1):
        cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        if col_idx == 5:  # Amount column
            cell.number_format = '#,##0'

# Column widths
ws_data.column_dimensions['A'].width = 20
ws_data.column_dimensions['B'].width = 15
ws_data.column_dimensions['C'].width = 25
ws_data.column_dimensions['D'].width = 15
ws_data.column_dimensions['E'].width = 15
ws_data.column_dimensions['F'].width = 15

# ==================== Q1: Sales by Country ====================
ws1 = wb.create_sheet("Q1")

ws1['A1'] = "Q1: Sales Performance by Country"
ws1['A1'].font = title_font
ws1.merge_cells('A1:F1')
ws1['A1'].alignment = Alignment(horizontal='center')

headers_q1 = ['Country', 'Total Revenue (€)', 'Total Boxes', 'Sales People', 'Avg Revenue per Box (€)', '% of Total Revenue']
for col_idx, header in enumerate(headers_q1, 1):
    cell = ws1.cell(row=3, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# Country analysis
country_sales = df.groupby('Country').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum',
    'Sales Person': 'nunique'
}).reset_index()
country_sales.columns = ['Country', 'Total Revenue', 'Total Boxes', 'Sales People']
country_sales = country_sales.sort_values('Total Revenue', ascending=False)

total_revenue = df['Amount'].sum()

row = 4
for _, data_row in country_sales.iterrows():
    ws1.cell(row=row, column=1, value=data_row['Country']).border = border
    ws1.cell(row=row, column=2, value=data_row['Total Revenue']).border = border
    ws1.cell(row=row, column=2).number_format = '#,##0'
    ws1.cell(row=row, column=3, value=data_row['Total Boxes']).border = border
    ws1.cell(row=row, column=4, value=data_row['Sales People']).border = border
    # Formula for avg revenue per box
    formula_e = f"=B{row}/C{row}"
    ws1.cell(row=row, column=5, value=formula_e).border = border
    # Formula for % of total
    formula_f = f"=B{row}/{total_revenue}"
    cell_f = ws1.cell(row=row, column=6, value=formula_f)
    cell_f.border = border
    cell_f.number_format = '0.0%'
    
    for col in range(1, 7):
        ws1.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    row += 1

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws1.column_dimensions[col].width = 20

# ==================== Q2: Top Products ====================
ws2 = wb.create_sheet("Q2")

ws2['A1'] = "Q2: Top 5 Products Analysis"
ws2['A1'].font = title_font
ws2.merge_cells('A1:E1')
ws2['A1'].alignment = Alignment(horizontal='center')

# Top 5 by Revenue
ws2['A3'] = "Top 5 Products by Revenue"
ws2['A3'].font = Font(bold=True, size=12)
ws2.merge_cells('A3:E3')

headers_q2a = ['Rank', 'Product', 'Total Revenue (€)', 'Total Boxes', '% of Total Revenue']
for col_idx, header in enumerate(headers_q2a, 1):
    cell = ws2.cell(row=4, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

product_rev = df.groupby('Product').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum'
}).reset_index()
product_rev = product_rev.sort_values('Amount', ascending=False).head(5)

row = 5
rank = 1
for _, data_row in product_rev.iterrows():
    ws2.cell(row=row, column=1, value=rank).border = border
    ws2.cell(row=row, column=2, value=data_row['Product']).border = border
    ws2.cell(row=row, column=3, value=data_row['Amount']).border = border
    ws2.cell(row=row, column=3).number_format = '#,##0'
    ws2.cell(row=row, column=4, value=data_row['Boxes Shipped']).border = border
    formula_pct = f"=C{row}/{total_revenue}"
    cell_pct = ws2.cell(row=row, column=5, value=formula_pct)
    cell_pct.border = border
    cell_pct.number_format = '0.0%'
    
    for col in range(1, 6):
        ws2.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    row += 1
    rank += 1

# Top 5 by Volume
ws2['A11'] = "Top 5 Products by Volume (Boxes)"
ws2['A11'].font = Font(bold=True, size=12)
ws2.merge_cells('A11:E11')

headers_q2b = ['Rank', 'Product', 'Total Boxes', 'Total Revenue (€)', 'Avg Revenue per Box']
for col_idx, header in enumerate(headers_q2b, 1):
    cell = ws2.cell(row=12, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

product_vol = df.groupby('Product').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum'
}).reset_index()
product_vol = product_vol.sort_values('Boxes Shipped', ascending=False).head(5)

row = 13
rank = 1
for _, data_row in product_vol.iterrows():
    ws2.cell(row=row, column=1, value=rank).border = border
    ws2.cell(row=row, column=2, value=data_row['Product']).border = border
    ws2.cell(row=row, column=3, value=data_row['Boxes Shipped']).border = border
    ws2.cell(row=row, column=4, value=data_row['Amount']).border = border
    ws2.cell(row=row, column=4).number_format = '#,##0'
    formula_avg = f"=D{row}/C{row}"
    ws2.cell(row=row, column=5, value=formula_avg).border = border
    
    for col in range(1, 6):
        ws2.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    row += 1
    rank += 1

for col in ['A', 'B', 'C', 'D', 'E']:
    ws2.column_dimensions[col].width = 22

# ==================== Q3: Sales Person Performance ====================
ws3 = wb.create_sheet("Q3")

ws3['A1'] = "Q3: Top 10 Sales People Performance"
ws3['A1'].font = title_font
ws3.merge_cells('A1:F1')
ws3['A1'].alignment = Alignment(horizontal='center')

headers_q3 = ['Rank', 'Sales Person', 'Total Revenue (€)', 'Total Boxes', 'Countries Covered', 'Products Sold']
for col_idx, header in enumerate(headers_q3, 1):
    cell = ws3.cell(row=3, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

sales_perf = df.groupby('Sales Person').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum',
    'Country': 'nunique',
    'Product': 'nunique'
}).reset_index()
sales_perf.columns = ['Sales Person', 'Total Revenue', 'Total Boxes', 'Countries', 'Products']
sales_perf = sales_perf.sort_values('Total Revenue', ascending=False).head(10)

row = 4
rank = 1
for _, data_row in sales_perf.iterrows():
    ws3.cell(row=row, column=1, value=rank).border = border
    ws3.cell(row=row, column=2, value=data_row['Sales Person']).border = border
    ws3.cell(row=row, column=3, value=data_row['Total Revenue']).border = border
    ws3.cell(row=row, column=3).number_format = '#,##0'
    ws3.cell(row=row, column=4, value=data_row['Total Boxes']).border = border
    ws3.cell(row=row, column=5, value=data_row['Countries']).border = border
    ws3.cell(row=row, column=6, value=data_row['Products']).border = border
    
    for col in range(1, 7):
        ws3.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    row += 1
    rank += 1

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws3.column_dimensions[col].width = 20

# ==================== Q4: Monthly Trends ====================
ws4 = wb.create_sheet("Q4")

ws4['A1'] = "Q4: Monthly Sales Trends (Jan-Aug 2022)"
ws4['A1'].font = title_font
ws4.merge_cells('A1:F1')
ws4['A1'].alignment = Alignment(horizontal='center')

headers_q4 = ['Month', 'Revenue (€)', 'Boxes Shipped', 'MoM Revenue Change', 'MoM Boxes Change', '% of Total Revenue']
for col_idx, header in enumerate(headers_q4, 1):
    cell = ws4.cell(row=3, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

monthly = df.groupby('Month').agg({
    'Amount': 'sum',
    'Boxes Shipped': 'sum'
}).reset_index()
monthly = monthly.sort_values('Month')

row = 4
prev_row = None
for _, data_row in monthly.iterrows():
    ws4.cell(row=row, column=1, value=data_row['Month']).border = border
    ws4.cell(row=row, column=2, value=data_row['Amount']).border = border
    ws4.cell(row=row, column=2).number_format = '#,##0'
    ws4.cell(row=row, column=3, value=data_row['Boxes Shipped']).border = border
    
    formula_pct = f"=B{row}/{total_revenue}"
    cell_pct = ws4.cell(row=row, column=6, value=formula_pct)
    cell_pct.border = border
    cell_pct.number_format = '0.0%'
    
    if prev_row:
        formula_mom_rev = f"=(B{row}-B{prev_row})/B{prev_row}"
        formula_mom_box = f"=(C{row}-C{prev_row})/C{prev_row}"
        cell_mom_rev = ws4.cell(row=row, column=4, value=formula_mom_rev)
        cell_mom_rev.border = border
        cell_mom_rev.number_format = '+0.0%;-0.0%'
        cell_mom_box = ws4.cell(row=row, column=5, value=formula_mom_box)
        cell_mom_box.border = border
        cell_mom_box.number_format = '+0.0%;-0.0%'
    else:
        ws4.cell(row=row, column=4, value="N/A").border = border
        ws4.cell(row=row, column=5, value="N/A").border = border
    
    for col in range(1, 7):
        ws4.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    prev_row = row
    row += 1

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws4.column_dimensions[col].width = 20

# ==================== Q5: Product-Country Matrix ====================
ws5 = wb.create_sheet("Q5")

ws5['A1'] = "Q5: Product Performance by Country (Revenue Matrix)"
ws5['A1'].font = title_font
ws5.merge_cells('A1:H1')
ws5['A1'].alignment = Alignment(horizontal='center')

# Create pivot matrix
matrix = df.pivot_table(values='Amount', index='Product', columns='Country', aggfunc='sum', fill_value=0)
matrix['Total'] = matrix.sum(axis=1)
matrix = matrix.sort_values('Total', ascending=False)

# Header
ws5.cell(row=3, column=1, value="Product").fill = header_fill
ws5.cell(row=3, column=1).font = header_font
ws5.cell(row=3, column=1).alignment = Alignment(horizontal='center')
ws5.cell(row=3, column=1).border = border

countries = list(matrix.columns)
for col_idx, country in enumerate(countries, 2):
    cell = ws5.cell(row=3, column=col_idx, value=country)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

# Data
row = 4
for product, data_row in matrix.iterrows():
    ws5.cell(row=row, column=1, value=product).border = border
    ws5.cell(row=row, column=1).alignment = Alignment(horizontal='left')
    
    for col_idx, country in enumerate(countries, 2):
        value = data_row[country]
        cell = ws5.cell(row=row, column=col_idx, value=value)
        cell.number_format = '#,##0'
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    row += 1

ws5.column_dimensions['A'].width = 25
for i in range(2, len(countries) + 2):
    ws5.column_dimensions[chr(64 + i)].width = 15

# Save workbook
output_path = 'C:/Users/Karen/.openclaw/workspace/Chocolate_Sales_Analysis.xlsx'
wb.save(output_path)
print(f"Excel file created: {output_path}")
print("\nSheets created:")
print("1. Chocolate Sales Data (original dataset)")
print("2. Q1: Sales by Country")
print("3. Q2: Top Products")
print("4. Q3: Sales Person Performance")
print("5. Q4: Monthly Trends")
print("6. Q5: Product-Country Matrix")
print("\nAll sheets have proper formatting and Excel formulas!")
